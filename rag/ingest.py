"""文档加载、切分与入库 CLI（增量入库，文件指纹驱动）。

支持格式：.txt .md .pdf .docx .xlsx .png .jpg .jpeg .bmp .tiff
用法：
    python -m rag.ingest               # 增量入库：仅处理新增/修改/删除的文件
    python -m rag.ingest --rebuild      # 清空重建索引（全量入库 + 重置清单）
"""

import argparse
import hashlib
import json
import logging
import sys
from pathlib import Path

from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from config.settings import get_settings
from rag.vectorstore import add_documents, add_image_documents, delete_by_ids, drop_collection

logger = logging.getLogger(__name__)

# 支持的文件扩展名
SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx",
                         ".png", ".jpg", ".jpeg", ".bmp", ".tiff"}

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff"}

# 文件清单文件名（记录每个文件的内容哈希与 chunk_ids，用于增量更新）
MANIFEST_FILENAME = ".ingest_manifest.json"


def load_text_file(filepath: Path) -> list[Document]:
    """加载 .txt/.md 文件。"""
    try:
        text = filepath.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = filepath.read_text(encoding="gbk", errors="ignore")
    return [Document(page_content=text, metadata={"source": str(filepath.name), "title": filepath.stem})]


def load_pdf(filepath: Path) -> list[Document]:
    """PDF 加载：文本优先，扫描页 OCR 兜底。

    1. 用 pypdf 提取文本
    2. 若某页文本过少（<配置阈值，疑似扫描件），渲染为图片 → OCR
    3. 对含图像的页面，额外生成图像 Document（metadata.image_path）
    """
    from rag.ocr import ocr_pdf, render_pdf_pages

    page_results = ocr_pdf(filepath)
    docs = []

    for pr in page_results:
        text = pr["text"]
        if text.strip():
            docs.append(Document(
                page_content=text,
                metadata={
                    "source": str(filepath.name),
                    "title": filepath.stem,
                    "page": pr["page"] + 1,
                    "ocr_used": pr["ocr_used"],
                },
            ))

    # 如果 PDF 有页面，额外为每页生成图像 Document（用于多模态检索）
    try:
        page_images = render_pdf_pages(filepath)
        import io

        image_dir = get_settings().docs_dir / ".page_images" / filepath.stem
        image_dir.mkdir(parents=True, exist_ok=True)
        for i, img in enumerate(page_images):
            img_path = image_dir / f"page_{i+1}.png"
            img.save(str(img_path), "PNG")
            docs.append(Document(
                page_content=f"[PDF页面图像 {filepath.name} 第{i+1}页]",
                metadata={
                    "source": str(filepath.name),
                    "title": filepath.stem,
                    "page": i + 1,
                    "image_path": str(img_path),
                    "type": "image",
                },
            ))
    except Exception as e:
        logger.warning("PDF 页面图像生成跳过: %s", e)

    return docs


def load_docx(filepath: Path) -> list[Document]:
    """Word .docx 加载：使用 langchain Docx2txtLoader。"""
    from langchain_community.document_loaders import Docx2txtLoader

    loader = Docx2txtLoader(str(filepath))
    docs = loader.load()
    # 补充 metadata
    for doc in docs:
        doc.metadata.setdefault("source", str(filepath.name))
        doc.metadata.setdefault("title", filepath.stem)
    return docs


def load_excel(filepath: Path) -> list[Document]:
    """Excel .xlsx 加载：openpyxl 遍历所有 sheet，提取单元格文本。"""
    from openpyxl import load_workbook

    wb = load_workbook(str(filepath), data_only=True)
    docs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c).strip()]
            if cells:
                rows_text.append(" | ".join(cells))
        if rows_text:
            text = "\n".join(rows_text)
            docs.append(Document(
                page_content=text,
                metadata={
                    "source": str(filepath.name),
                    "title": filepath.stem,
                    "sheet": sheet_name,
                },
            ))
    wb.close()
    return docs


def load_image(filepath: Path) -> list[Document]:
    """图片加载：OCR 提取文本 + 图像 Document（用于多模态检索）。"""
    from rag.ocr import ocr_image

    docs = []
    # OCR 提取文本
    try:
        ocr_text = ocr_image(filepath)
        if ocr_text.strip():
            docs.append(Document(
                page_content=ocr_text,
                metadata={
                    "source": str(filepath.name),
                    "title": filepath.stem,
                    "ocr": True,
                },
            ))
    except Exception as e:
        logger.warning("图片 OCR 失败 %s: %s", filepath.name, e)

    # 图像 Document（page_content 用占位描述，metadata.image_path 存路径）
    docs.append(Document(
        page_content=f"[图片 {filepath.name}]",
        metadata={
            "source": str(filepath.name),
            "title": filepath.stem,
            "image_path": str(filepath),
            "type": "image",
        },
    ))
    return docs


def load_file(filepath: Path) -> list[Document]:
    """按扩展名分发到对应加载器。"""
    ext = filepath.suffix.lower()
    if ext in (".txt", ".md"):
        return load_text_file(filepath)
    elif ext == ".pdf":
        return load_pdf(filepath)
    elif ext == ".docx":
        return load_docx(filepath)
    elif ext == ".xlsx":
        return load_excel(filepath)
    elif ext in IMAGE_EXTENSIONS:
        return load_image(filepath)
    else:
        logger.warning("不支持的文件格式: %s", filepath)
        return []


def load_files(docs_dir: Path) -> list[Document]:
    """加载 docs_dir 下所有支持格式的文件为 Document（全量加载，用于 rebuild 场景）。"""
    docs: list[Document] = []
    for fp in sorted(docs_dir.iterdir()):
        if fp.is_file() and fp.suffix.lower() in SUPPORTED_EXTENSIONS:
            try:
                docs.extend(load_file(fp))
                logger.info("已加载: %s", fp.name)
            except Exception as e:
                logger.error("加载失败 %s: %s", fp.name, e)
    return docs


def split_documents(docs: list[Document]) -> list[Document]:
    """使用递归字符切分器将文档切分为小块。

    切片大小与重叠从配置读取（rag_chunk_size / rag_chunk_overlap）。
    图像 Document（metadata.type=="image"）不参与切分，直接保留。
    """
    settings = get_settings()
    text_docs = [d for d in docs if d.metadata.get("type") != "image"]
    image_docs = [d for d in docs if d.metadata.get("type") == "image"]

    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.rag_chunk_size,
        chunk_overlap=settings.rag_chunk_overlap,
        separators=["\n\n", "\n", "。", "！", "？", "；", ".", " ", ""],
    )
    chunks = splitter.split_documents(text_docs)
    # 图像文档直接追加
    chunks.extend(image_docs)
    return chunks


# ===== 文件清单（manifest）管理：记录每个文件的内容哈希与 chunk_ids，支撑增量更新 =====


def _manifest_path() -> Path:
    """返回清单文件绝对路径（与 docs_dir 同目录下的隐藏 JSON）。"""
    return get_settings().docs_dir / MANIFEST_FILENAME


def compute_file_hash(filepath: Path) -> str:
    """计算文件内容 md5 哈希（按块读取，支持大文件）。"""
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for block in iter(lambda: f.read(8192), b""):
            h.update(block)
    return h.hexdigest()


def load_manifest() -> dict:
    """读取文件清单。清单记录 {文件名: {"hash", "mtime", "chunk_ids"}}。"""
    p = _manifest_path()
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        logger.warning("读取 manifest 失败，视为空: %s", e)
        return {}


def save_manifest(manifest: dict) -> None:
    """写入文件清单（UTF-8，便于人工排查）。"""
    p = _manifest_path()
    try:
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as e:
        logger.warning("写入 manifest 失败: %s", e)


def ingest_one_file(filepath: Path) -> list[str]:
    """加载、切分并入库单个文件，返回新增 chunk 的 id 列表。

    文本块与图像块分别入库（图像块走多模态 Embedding 路径）。
    """
    docs = load_file(filepath)
    if not docs:
        return []
    chunks = split_documents(docs)
    text_chunks = [c for c in chunks if c.metadata.get("type") != "image"]
    image_chunks = [c for c in chunks if c.metadata.get("type") == "image"]
    ids: list[str] = []
    if text_chunks:
        ids.extend(add_documents(text_chunks))
    if image_chunks:
        ids.extend(add_image_documents(image_chunks))
    return ids


def _scan_current_files() -> dict:
    """扫描 docs_dir 下当前支持格式的文件，返回 {文件名: Path}。"""
    docs_dir = get_settings().docs_dir
    files = {}
    if not docs_dir.exists():
        return files
    for fp in sorted(docs_dir.iterdir()):
        if fp.is_file() and fp.suffix.lower() in SUPPORTED_EXTENSIONS:
            files[fp.name] = fp
    return files


def ingest(rebuild: bool = False) -> int:
    """增量入库：仅处理新增/修改/删除的文件，返回本次新增（含更新）的 chunk 数。

    流程：
      1. rebuild=True：drop_collection 清空索引 + 清空清单 → 全部文件视为新增
      2. 扫描当前文件，对每个文件先比 mtime 再比 hash 判断是否变化
      3. 新增/修改文件：（修改时先删旧 chunk）→ 重新入库 → 记录 chunk_ids
      4. 未变化文件：跳过，沿用旧 chunk_ids
      5. 清单中有但目录无（已删除文件）：按 chunk_ids 清理孤儿 chunk
      6. 写回清单
    """
    settings = get_settings()
    manifest = {}

    if rebuild:
        try:
            # 删除并重建 Milvus collection（清空旧索引）
            drop_collection()
            print("[ingest] 已清空旧索引，开始全量重建")
        except Exception as e:
            print(f"[ingest] 清空索引跳过: {e}")
        manifest = {}
    else:
        manifest = load_manifest()

    current_files = _scan_current_files()
    if not current_files:
        print(f"[ingest] 未在 {settings.docs_dir} 找到支持的文档")
        print(f"[ingest] 支持格式: {', '.join(sorted(SUPPORTED_EXTENSIONS))}")
        # 即使无文档也写入空清单，避免残留
        save_manifest({})
        return 0

    new_manifest: dict = {}
    added_files = updated_files = skipped_files = deleted_files = 0
    total_added = 0

    # 处理新增 / 修改文件
    for fname, fp in current_files.items():
        old = manifest.get(fname)
        mtime = fp.stat().st_mtime
        # mtime 预筛：未变则直接沿用，跳过 hash 计算（快速路径）
        if old and old.get("mtime") == mtime:
            new_manifest[fname] = old
            skipped_files += 1
            continue
        # mtime 变了，算 hash 确认内容是否真改
        fhash = compute_file_hash(fp)
        if old and old.get("hash") == fhash:
            # mtime 变但内容未变（如 touch），沿用并更新 mtime
            new_manifest[fname] = {**old, "mtime": mtime}
            skipped_files += 1
            continue

        # 内容确实变化（新增或修改）
        if old:
            # 修改文件：先删除该文件旧 chunk，避免重复数据
            try:
                delete_by_ids(old.get("chunk_ids", []))
            except Exception as e:
                logger.warning("删除旧 chunk 失败 %s: %s", fname, e)

        try:
            chunk_ids = ingest_one_file(fp)
        except Exception as e:
            logger.error("入库失败 %s: %s", fname, e)
            # 失败时若原为修改文件，保留旧记录以便下次重试；新文件则不记录
            if old:
                new_manifest[fname] = old
            continue

        new_manifest[fname] = {
            "hash": fhash,
            "mtime": mtime,
            "chunk_ids": chunk_ids,
        }
        total_added += len(chunk_ids)
        if old:
            updated_files += 1
            logger.info("已更新: %s (%d chunks)", fname, len(chunk_ids))
        else:
            added_files += 1
            logger.info("已入库: %s (%d chunks)", fname, len(chunk_ids))

    # 处理已删除文件（清单有但目录无）：清理孤儿 chunk
    for fname, old in manifest.items():
        if fname not in current_files:
            try:
                delete_by_ids(old.get("chunk_ids", []))
                deleted_files += 1
                logger.info("已清理删除文件: %s", fname)
            except Exception as e:
                logger.warning("清理孤儿 chunk 失败 %s: %s", fname, e)

    save_manifest(new_manifest)
    print(
        f"[ingest] 新增 {added_files}，更新 {updated_files}，跳过 {skipped_files}，"
        f"删除 {deleted_files} 个文件；本次入库 {total_added} chunks"
    )
    return total_added


def main():
    parser = argparse.ArgumentParser(description="RAG 多模态文档入库工具（增量）")
    parser.add_argument("--rebuild", action="store_true", help="清空旧索引后全量重建")
    args = parser.parse_args()
    count = ingest(rebuild=args.rebuild)
    print(f"[ingest] 完成，本次处理 {count} 个 chunk")
    return count


if __name__ == "__main__":
    main()
